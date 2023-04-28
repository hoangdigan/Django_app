from django.shortcuts import render

import pandas as pd
from django.http.response import HttpResponse
import mimetypes
import random

import os
import glob

from datetime import datetime
import pandas as pd

import mysql.connector
from mysql.connector import Error
import openpyxl
import openpyxl.utils.cell 

from openpyxl import Workbook

from openpyxl.chart import PieChart, Reference, Series

def index(request):
    
    dataprocess()
    return render(request, "index.html")      

def industry():
    pass

def macro():
    pass

def ratings():
    pass
def other():
    pass

def dataprocess():    
    try:
        conn = mysql.connector.connect(host='localhost',
                            database='saigonrates',
                            user='root',
                            password='123456')

        cursor = conn.cursor()

        sql =   """select * from financial_statement""" 
        cursor.execute(sql)       
        df = pd.DataFrame.from_records(cursor.fetchall(),
                                columns = [desc[0] for desc in cursor.description])
        print("df", df)   
        df.to_excel("fs_test.xlsx")    

        workbook = openpyxl.load_workbook("fs_test.xlsx") 
        sheet = workbook.active
        maxrow = sheet.max_row 

        for row in sheet.iter_rows(min_row=1, min_col=4, max_row=94, max_col=4, values_only=False):
            for cell in row:
                if cell.value =="117":
                    doanh_so=cell.row
                    
                if cell.value =="125":
                    cpbh=cell.row

                if cell.value =="126":
                    cpqldn=cell.row
                    break               
        c= openpyxl.utils.cell.get_column_letter(8)       
        sheet.cell(maxrow+2, column=6).value = "Selling & GA expense/ Revenue"
        sheet.cell(maxrow+2, column=8).value = "=({c1}{r1} + {c1}{r2}) /{c1}{r3}".format(c1=c, r1 = cpbh, r2=cpqldn, r3 =doanh_so)

        for r in range(2,95):
            sheet[f'H{r}'].number_format ='"$"#,##0_);("$"#,##0)'    
        
        # ws = sheet
        # wb = workbook
        cs = workbook.create_chartsheet()

        rows = [
            ["Bob", 3],
            ["Harry", 2],
            ["James", 4],
        ]

        # for row in rows:
        #     sheet.append(row)
        sheet.cell(100, column=1).value ="Bob"
        sheet.cell(101, column=1).value ="Harry"
        sheet.cell(102, column=1).value ="James"
        sheet.cell(100, column=2).value =2
        sheet.cell(101, column=2).value =3
        sheet.cell(102, column=2).value =4

        chart = PieChart()
        labels = Reference(sheet, min_col=1, min_row=100, max_row=102)
        data = Reference(sheet, min_col=2, min_row=100, max_row=102)
        chart.series = (Series(data),)
        chart.title = "PieChart"

        cs.add_chart(chart)        
            
        workbook.save('realpython.xlsx')       


    except Error as e:
        print(e)

    finally:
        cursor.close()
        conn.close()

def fs_process():
    wb1 = openpyxl.load_workbook("n_proccess.xlsx") 
    ws1 = wb1.active

    wb2 = openpyxl.load_workbook("fsa_dn_template.xlsx") 
    ws2 = wb2['BS.data']
    ws3 = wb2['PL.data']
    ws4 = wb2['CF.data']

    max_row = ws1.max_row 
    year_list = [2004,2005,2006,2007,2008,2009,2010,2011,2012,2013,2014,2015,2016,2017,2018,2019,2020,2021,2022]

    j=5
    for year in year_list:
        for i in range(2, max_row+1):
            if i==year:
                if ws1.cells(i,6) == "BCDKT":
                    for k in range(4,120):
                        if ws2.cells(k,2) == ws1.cell(i,1):
                            ws2.cells(k,j).value = ws1.cell(i,5).value
        j+=1


    

def download_file(request):
    
    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    # Define text file name
    filename = 'realpython.xlsx'
    # Define the full file path
    filepath = os.path.join(BASE_DIR,'download',filename)
    # Open the file for reading content
    path = open(filepath, "rb")
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    # response = HttpResponse(path, content_type=mime_type)
    response = HttpResponse(path, content_type="application/ms-excel")
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename
    # Return the response value
    return response


    

    
